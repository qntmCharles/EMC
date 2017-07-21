import statistics, math, openpyxl
import numpy as np
from datetime import datetime
from openpyxl import Workbook

basedir = '/home/cwp/EMC/lib/analysis/variation/temporal/'

def get(path):
    data = []
    with open(path,'r') as f:
        lines = f.readlines()
        for line in lines:
           data.append(line.split('\n')[0])

    return data

def day(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for day in range(1, 32):
        print(day)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.day == day):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(day)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(day)

    finalData = np.ma.masked_object(plotData, None)
    finalErr = np.ma.masked_object(plotErr, None)

    return plotTimes, finalData, finalErr

def month(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for month in range(1,13):
        print(month)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.month == month):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(month)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(month)

    finalData = np.ma.masked_object(plotData, None)
    finalErr = np.ma.masked_object(plotErr, None)

    return finalData, finalErr, plotTimes

def year(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for year in range(2000,2017):
        print(year)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.year == year):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(year)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(year)

    finalData = np.ma.masked_object(plotData, None)
    finalErr = np.ma.masked_object(plotErr, None)

    return finalData, finalErr, plotTimes

def yearmonth(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for year in range(2000,2017):
        print(year)
        for month in range(1,13):
            print(month)
            currentData = []

            for i in range(len(datas)):
                currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
                if (currentDate.year == year) and (currentDate.month == month):
                    currentData.extend([int(x) for x in datas[i].split(',')])

            if len(currentData) != 0:
                plotTimes.append(datetime.strptime(str(year)+'-'+'{0:02d}'.format(month), "%Y-%m"))
                plotData.append(statistics.mean(currentData))
                if len(currentData) > 1:
                    plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
                else:
                    plotErr.append(1)
            else:
                plotErr.append(None)
                plotData.append(None)
                plotTimes.append(datetime.strptime(str(year)+'-'+'{0:02d}'.format(month), "%Y-%m"))

    finalData = np.ma.masked_object(plotData, None)
    finalErr = np.ma.masked_object(plotErr, None)

    return finalData, finalErr, plotTimes

Ndata_full = get(basedir+'NS/NplotData.txt')
Ntimes_full = get(basedir+'NS/NplotTimes.txt')

data_full = get('/home/cwp/EMC/lib/analysis/plotData.txt')
times_full = get('/home/cwp/EMC/lib/analysis/plotTimes.txt')

Nday_times, Nday_data, Nday_err = day(Ndata_full, Ntimes_full)
day_times, day_data, day_err = day(data_full, times_full)

Nmonth_data, Nmonth_err, Nmonth_times = month(Ndata_full, Ntimes_full)
month_data, month_err, month_times = month(data_full, times_full)

Nyear_data, Nyear_err, Nyear_times = year(Ndata_full,Ntimes_full)
year_data, year_err, year_times = year(data_full,times_full)

Nyearmonth_data, Nyearmonth_err, Nyearmonth_times = yearmonth(Ndata_full,\
        Ntimes_full)
yearmonth_data, yearmonth_err, yearmonth_times = yearmonth(data_full,\
        times_full)

wb = Workbook()

savefile = 'data.xlsx'

ws1 = wb.active
ws1.title = "Full data"
ws2 = wb.create_sheet(title="Day scale")
ws3 = wb.create_sheet(title="Month scale")
ws4 = wb.create_sheet(title="Year scale")
ws5 = wb.create_sheet(title="Year-month scale")

def enterIntoWorkbook(wb, ws, lists, titles, startCol):
    for i in range(startCol,startCol+len(lists)):
        ws.cell(column=i,row=1, value=titles[i-startCol])
        for j in range(1,len(lists[i-startCol])+1):
            ws.cell(column=i,row=j+1,value=lists[i-startCol][j-1])

    wb.save(filename = savefile)

enterIntoWorkbook(wb,ws1,[times_full,data_full],["Time", "Data"],1)
enterIntoWorkbook(wb,ws1,[Ntimes_full,Ndata_full],["N_Time", "N_Data"],4)

enterIntoWorkbook(wb,ws2,[day_times,day_data,day_err],["Time","Data","SE"],1)
enterIntoWorkbook(wb,ws2,[Nday_times,Nday_data,Nday_err],["N_Time","N_Data","N_SE"],4)

enterIntoWorkbook(wb,ws3,[month_times,month_data,month_err],["Time","Data","SE"],1)
enterIntoWorkbook(wb,ws3,[Nmonth_times,Nmonth_data,Nmonth_err],["N_Time","N_Data","N_SE"],4)

enterIntoWorkbook(wb,ws4,[year_times,year_data,year_err],["Time","Data","SE"],1)
enterIntoWorkbook(wb,ws4,[Nyear_times,Nyear_data,Nyear_err],["N_Time","N_Data","N_SE"],4)

enterIntoWorkbook(wb,ws5,[yearmonth_times,yearmonth_data,yearmonth_err],["Time","Data","SE"],1)
enterIntoWorkbook(wb,ws5,[Nyearmonth_times,Nyearmonth_data,Nyearmonth_err],["N_Time","N_Data","N_SE"],4)
