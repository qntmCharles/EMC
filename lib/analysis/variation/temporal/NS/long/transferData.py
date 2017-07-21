import statistics, math, openpyxl
import numpy as np
from datetime import datetime
from openpyxl import Workbook

basedir = '/home/cwp/EMC/lib/analysis/variation/temporal/NS/long/'

def get(path):
    data = []
    with open(path,'r') as f:
        lines = f.readlines()
        for line in lines:
           data.append(line.split('\n')[0])

    return data

def day(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for day in range(1, 32):
        print(day)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.day == day):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(day)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(day)

    #finalData = np.ma.masked_object(plotData, None)
    #finalErr = np.ma.masked_object(plotErr, None)

    return plotTimes, plotData, plotErr

def month(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for month in range(1,13):
        print(month)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.month == month):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(month)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(month)

    #finalData = np.ma.masked_object(plotData, None)
    #finalErr = np.ma.masked_object(plotErr, None)

    return plotData, plotErr, plotTimes

def year(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for year in range(2000,2017):
        print(year)
        currentData = []

        for i in range(len(datas)):
            currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
            if (currentDate.year == year):
                currentData.extend([int(x) for x in datas[i].split(',')])

        if len(currentData) != 0:
            plotTimes.append(year)
            plotData.append(statistics.mean(currentData))
            if len(currentData) > 1:
                plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
            else:
                plotErr.append(1)
        else:
            plotErr.append(None)
            plotData.append(None)
            plotTimes.append(year)

    #finalData = np.ma.masked_object(plotData, None)
    #finalErr = np.ma.masked_object(plotErr, None)

    return plotData, plotErr, plotTimes

def yearmonth(datas, times):
    plotTimes = []
    plotData = []
    plotErr = []

    for year in range(2000,2017):
        print(year)
        for month in range(1,13):
            print(month)
            currentData = []

            for i in range(len(datas)):
                currentDate = datetime.strptime(times[i], "%Y-%m-%d %H:%M:%S")
                if (currentDate.year == year) and (currentDate.month == month):
                    currentData.extend([int(x) for x in datas[i].split(',')])

            if len(currentData) != 0:
                plotTimes.append(datetime.strptime(str(year)+'-'+'{0:02d}'.format(month), "%Y-%m"))
                plotData.append(statistics.mean(currentData))
                if len(currentData) > 1:
                    plotErr.append(statistics.stdev(currentData)/math.sqrt(len(currentData)))
                else:
                    plotErr.append(1)
            else:
                plotErr.append(None)
                plotData.append(None)
                plotTimes.append(datetime.strptime(str(year)+'-'+'{0:02d}'.format(month), "%Y-%m"))

    #finalData = np.ma.masked_object(plotData, None)
    #finalErr = np.ma.masked_object(plotErr, None)

    return plotData, plotErr, plotTimes


wb = Workbook()

savefile = 'long.xlsx'

ws1 = wb.active
ws1.title = "Full data"
ws2 = wb.create_sheet(title="Day scale")
ws3 = wb.create_sheet(title="Month scale")
ws4 = wb.create_sheet(title="Year scale")
ws5 = wb.create_sheet(title="Year-month scale")

def enterIntoWorkbook(wb, ws, lists, titles, startCol):
    for i in range(startCol,startCol+len(lists)):
        ws.cell(column=i,row=1, value=titles[i-startCol])
        for j in range(1,len(lists[i-startCol])+1):
            ws.cell(column=i,row=j+1,value=lists[i-startCol][j-1])

    wb.save(filename = savefile)

count=1
count2=1
for category in ["0to15","15to30","30to45","45to60","105to120","135to150",\
        "150to165","0toM15","M30toM45","M45toM60","M60toM75","M75toM90",\
        "M90toM105","M105toM120","M120toM135"]:
    for DN in ['D', 'N']:
        data_full = get(basedir+category+DN+'plotData.txt')
        times_full = get(basedir+category+DN+'plotTimes.txt')

        day_times, day_data, day_err = day(data_full, times_full)
        month_data, month_err, month_times = month(data_full, times_full)
        year_data, year_err, year_times = year(data_full,times_full)
        yearmonth_data, yearmonth_err, yearmonth_times = yearmonth(data_full,\
                times_full)

        enterIntoWorkbook(wb,ws1,[times_full,data_full],[category+DN+"Time", category+DN+"Data"],count2)
        enterIntoWorkbook(wb,ws2,[day_times,day_data,day_err],[category+DN+"Time",category+DN+"Data",category+DN+"SE"],count)
        enterIntoWorkbook(wb,ws3,[month_times,month_data,month_err],[category+DN+"Time",category+DN+"Data",category+DN+"SE"],count)
        enterIntoWorkbook(wb,ws4,[year_times,year_data,year_err],[category+DN+"Time",category+DN+"Data",category+DN+"SE"],count)
        enterIntoWorkbook(wb,ws5,[yearmonth_times,yearmonth_data,yearmonth_err],[category+DN+"Time",category+DN+"Data",category+DN+"SE"],count)
        count+=3
        count2+=2
